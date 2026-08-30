"""
Bulk-imports leads from a 4-column .xlsx file (Name, Organisation,
Profile/Designation, LinkedIn) into the unified `leads` MongoDB collection.
Columns are matched by intent, not exact text, and may appear in any order
(see `detect_columns` below) so minor header variations ("Company",
"Designation", "LinkedIn URL", etc.) don't break the import.

Setup (once per machine):
    pip3 install -r scripts/requirements.txt

Usage:
    python3 scripts/import_excel.py <path-to-xlsx-file>

Prints exactly one JSON line to stdout on completion:
    {"success": true, "totalRows": N, "inserted": N, "updated": N, "skipped": N}
or on failure:
    {"success": false, "error": "..."}

The dedupeKey normalization below MUST stay identical to the one in
scripts/migrateToLeads.ts so both writers upsert into the same rows.
"""

import json
import os
import re
import sys
from datetime import datetime, timezone

from dotenv import load_dotenv, find_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne

# Order matters: checked top-to-bottom, first match wins per header cell.
# Keep "organisation" checks before "name" isn't needed since none overlap,
# but linkedin/organisation are checked before name/profile to avoid a stray
# "name" substring (e.g. "Company Name") being misread as the name column.
COLUMN_MATCHERS = [
    ("linkedin", ["linkedin", "li url", "profile url", "profile link"]),
    ("organisation", ["organisation", "organization", "company", "employer", "firm"]),
    ("role", ["profile", "designation", "role", "title", "position"]),
    ("name", ["name"]),
]


def fail(message: str, exit_code: int = 1):
    print(json.dumps({"success": False, "error": message}))
    sys.exit(exit_code)


def normalize_header(header) -> str:
    text = str(header).strip().lower() if header is not None else ""
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def detect_columns(header_row):
    """Maps each required field ('name', 'organisation', 'role', 'linkedin')
    to a column index, matching headers by keyword rather than exact text.
    Returns (mapping, missing_fields)."""
    normalized = [normalize_header(h) for h in header_row]
    mapping = {}

    for field, keywords in COLUMN_MATCHERS:
        for idx, header in enumerate(normalized):
            if idx in mapping.values():
                continue
            if any(keyword in header for keyword in keywords):
                mapping[field] = idx
                break

    required = {"name", "organisation"}
    missing = required - mapping.keys()
    return mapping, missing


def normalize_linkedin(url: str) -> str:
    value = url.strip().lower()
    if value.startswith("https://"):
        value = value[len("https://"):]
    elif value.startswith("http://"):
        value = value[len("http://"):]
    if value.startswith("www."):
        value = value[len("www."):]
    value = value.split("?")[0]
    while value.endswith("/"):
        value = value[:-1]
    return value


def make_dedupe_key(name: str, organisation: str, linkedin):
    if linkedin and linkedin.strip():
        return normalize_linkedin(linkedin)
    return f"{name.strip().lower()}|{organisation.strip().lower()}"


def main():
    if len(sys.argv) != 2:
        fail("Usage: python3 import_excel.py <path-to-xlsx-file>")

    xlsx_path = sys.argv[1]
    if not os.path.isfile(xlsx_path):
        fail(f"File not found: {xlsx_path}")

    load_dotenv(find_dotenv(usecwd=True))
    uri = os.environ.get("MONGODB_URI")
    db_name = os.environ.get("MONGODB_NAME")
    if not uri or not db_name:
        fail("MONGODB_URI / MONGODB_NAME not set in environment or .env")

    try:
        workbook = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        fail(f"Could not open xlsx file: {exc}")

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        fail("Spreadsheet is empty")
        return

    mapping, missing = detect_columns(header_row)
    if missing:
        fail(
            "Could not find required column(s) "
            f"{sorted(missing)} in header row {header_row}. "
            "Expected columns for Name, Organisation, Profile/Designation, and LinkedIn "
            "(headers are matched by keyword, so close variants are fine, "
            "but Name and Organisation columns must be identifiable).",
            exit_code=2,
        )

    def cell(row, field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    total_rows = 0
    skipped = 0
    ops = []
    now = datetime.now(timezone.utc)

    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        total_rows += 1

        name = cell(row, "name")
        organisation = cell(row, "organisation")
        profile = cell(row, "role") or None
        linkedin = cell(row, "linkedin") or None

        if not name or not organisation:
            skipped += 1
            continue

        dedupe_key = make_dedupe_key(name, organisation, linkedin)
        doc = {
            "name": name,
            "organisation": organisation,
            "role": profile,
            "linkedin": linkedin,
            "email": None,
            "sector": None,
            "connectedOn": None,
            "source": "excel_import",
            "dedupeKey": dedupe_key,
            "updatedAt": now,
        }

        ops.append(
            UpdateOne(
                {"dedupeKey": dedupe_key},
                {"$set": doc, "$setOnInsert": {"createdAt": now}},
                upsert=True,
            )
        )

    inserted = 0
    updated = 0

    if ops:
        try:
            client = MongoClient(uri, serverSelectionTimeoutMS=10000)
            db = client[db_name]
            result = db["leads"].bulk_write(ops, ordered=False)
            inserted = result.upserted_count
            updated = result.modified_count
            client.close()
        except Exception as exc:
            fail(f"Database error: {exc}")

    print(
        json.dumps(
            {
                "success": True,
                "totalRows": total_rows,
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
            }
        )
    )
    sys.exit(0)


if __name__ == "__main__":
    main()
